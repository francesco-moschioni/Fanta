"""Reconstruct the G3/G4 (+ post-auction) assignments into the ledger from the
league-platform roster export, so the ledger reflects the real pre-riparazione
state (today it stops at G1+G2).

Source: "20lega-rosters-*.xlsx" (see src/fantacalcio/ingest/lega_rosters.py) +
the end-of-market listone "Quotazioni_Fantacalcio_Stagione_2026_27.xlsx".

Method (ADR-2026-081)
---------------------
1. Resolve every platform team name to a `team_id` via `ledger.team_labels`
   (19/20 exact; `team_05` renamed "Werder Bremer" -> "I Have a N'Drim").
2. Resolve every player display name to a `player_code` against the listone
   `Tutti` + `Ceduti` sheets, plus a 2-code fallback for old goalkeepers and an
   alias for the 4 synthetic negative codes (`-1..-4` -> real, ADR-2026-051/064).
   All 460 names resolve.
3. `owned` per team = the ledger's G1+G2 player_ids (negative-aliased).
   `new` = resolved roster codes - owned.
4. Every `new` non-goalkeeper line -> `AssignmentEvent` in round **G3**, pool
   `remaining_players` (open pool, any role). G3 and G4 share the pool and one
   carried budget (`remaining_G2 + 40`), so labelling everything G3 changes no
   invariant.
5. Goalkeeper blocks: 13/20 teams' roster keepers already match the G1 block. The
   other 7 swapped one keeper in G3/G4 -> a `VoidEvent` on the G1 block + a new
   3-keeper `AssignmentEvent` in G3 (pool `remaining_players`, role GK). Under
   the *current-view* replay (`effective_events`) the voided block is dropped, so
   the new block applies cleanly; the full audit replay would double it, which is
   why validation here uses `replay(effective_events(...))`.
6. Where a team's real roster cost exceeds its config-derived G3 budget
   (`remaining_G2 + 40`) -- the 340-vs-369 discrepancy, admin gave more credits
   than the base rounds imply -- a `BudgetAdjustmentEvent` in G3 closes the gap
   (`reason="lega_roster_reconciliation"`), so the honest spend replays and the
   discrepancy is on the record for the still-open admin-rules question.
7. Validate the whole thing via `replay(effective_events(existing + new))`.
   Dry-run by default; `--yes` appends; `--relabel-team05` also fixes the label.

Usage:
    python scripts/import_lega_rosters.py <roster.xlsx> <listone.xlsx> [--yes] [--relabel-team05]
"""

from __future__ import annotations

import argparse
import collections
import difflib
import sys
import unicodedata
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from fantacalcio.config import ConfigError, load_ruleset
from fantacalcio.domain import (
    AssignmentEvent,
    AssignmentItem,
    BudgetAdjustmentEvent,
    DomainError,
    Role,
    VoidEvent,
    effective_events,
    evaluate_budget_expr,
    replay,
)
from fantacalcio.ingest.lega_rosters import parse_roster_file
from fantacalcio.persistence import ledger_store, team_labels_store

ROUND_ID = "G3"
POOL_ID = "remaining_players"
GK_POOL_ID = "goalkeeper_blocks"
SOURCE = "lega_roster_export_import"
AUTHOR = "utente"

NEGATIVE_ALIAS = {"-1": "4998", "-2": "7329", "-3": "5982", "-4": "7551"}
OLD_GK_FALLBACK = {"lolic": ("7466", "P"), "satalino": ("2127", "P")}
TEAM05_OLD_LABEL = "Werder Bremer"
TEAM05_NEW_LABEL = "I Have a N'Drim"

_ROLE_MAP = {"P": Role.GK, "D": Role.DEF, "C": Role.MID, "A": Role.FWD}
_TS = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _positional_roles(n_slots: int, ruleset) -> list[Role]:
    """The roster grid is positional: the first `goalkeeper_block_size` rows are
    the keeper block, then `defenders`, then `midfielders`, then whatever is left
    are the forward slots (the league runs the `forwards_fallback` 4-forward
    roster, not the nominal 5). Returns one Role per slot, in grid order."""
    r = ruleset.roster
    gk, d, m = r.goalkeeper_block_size, r.defenders, r.midfielders
    fwd = n_slots - gk - d - m
    if fwd < 1:
        raise SystemExit(f"roster block sizes don't fit {n_slots} slots (gk{gk} d{d} m{m})")
    return [Role.GK] * gk + [Role.DEF] * d + [Role.MID] * m + [Role.FWD] * fwd


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower().replace(".", "").replace("'", " ").replace("-", " ").split("*")[0].strip()


def _squash(s: str) -> str:
    return "".join(ch for ch in _norm(s) if ch.isalnum())


def _load_listone_index(listone_xlsx: Path) -> dict[str, list[tuple[str, str]]]:
    wb = openpyxl.load_workbook(listone_xlsx, read_only=True, data_only=True)
    idx: dict[str, list[tuple[str, str]]] = collections.defaultdict(list)
    for sheet in ("Tutti", "Ceduti"):
        if sheet not in wb.sheetnames:
            continue
        for r in wb[sheet].iter_rows(min_row=3, values_only=True):
            if r[0] is None:
                continue
            try:
                code = str(int(r[0]))
            except (TypeError, ValueError):
                continue
            idx[_norm(r[3])].append((code, str(r[1])))
    wb.close()
    return idx


def _resolve_player(name: str, idx) -> tuple[str, str] | None:
    """-> (player_code, role_letter) or None."""
    n = _norm(name)
    if n in idx and len(idx[n]) == 1:
        code, role = idx[n][0]
        return NEGATIVE_ALIAS.get(code, code), role
    if n in OLD_GK_FALLBACK:
        return OLD_GK_FALLBACK[n]
    close = difflib.get_close_matches(n, list(idx), n=2, cutoff=0.86)
    if len(close) == 1 and len(idx[close[0]]) == 1:
        code, role = idx[close[0]][0]
        return NEGATIVE_ALIAS.get(code, code), role
    return None


def _resolve_team_id(name: str, labels: dict[str, str]) -> str | None:
    for tid, lab in labels.items():
        if _squash(lab) == _squash(name):
            return tid
    if _squash(name) == _squash(TEAM05_NEW_LABEL):
        for tid, lab in labels.items():
            if _squash(lab) == _squash(TEAM05_OLD_LABEL):
                return tid
    best, best_r = None, 0.0
    for tid, lab in labels.items():
        r = difflib.SequenceMatcher(None, _norm(name), _norm(lab)).ratio()
        if r > best_r:
            best, best_r = tid, r
    return best if best_r >= 0.80 else None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("roster_xlsx", type=Path)
    ap.add_argument("listone_xlsx", type=Path)
    ap.add_argument("--yes", action="store_true", help="append to the ledger (default: dry-run)")
    ap.add_argument("--relabel-team05", action="store_true")
    ap.add_argument(
        "--only-team",
        default=None,
        help="reconcile just this team_id (e.g. team_01). Some teams need the "
        "positional-role / ReleaseEvent handling not yet in the domain model "
        "(ADR-2026-073/081); --only-team lets the clean ones land now.",
    )
    args = ap.parse_args()

    ruleset = load_ruleset(Path("config/auction_rules.v1.yaml"))
    staged = parse_roster_file(args.roster_xlsx)
    idx = _load_listone_index(args.listone_xlsx)

    labels_conn = team_labels_store.connect()
    labels = team_labels_store.get_all_labels(labels_conn)
    ledger_conn = ledger_store.connect()
    existing = ledger_store.load_events(ledger_conn)

    # current-view state: owned codes per team (neg-aliased), G1 GK block event id,
    # and remaining_G2 per team for the budget-gap check.
    eff_state = replay(ruleset, effective_events(existing))
    owned: dict[str, set[str]] = collections.defaultdict(set)
    gk_block_event: dict[str, tuple[str, int]] = {}
    orig_event_by_code: dict[str, dict[str, AssignmentEvent]] = collections.defaultdict(dict)
    for ev in existing:
        if isinstance(ev, AssignmentEvent):
            for pid in ev.item.player_ids:
                aliased = NEGATIVE_ALIAS.get(pid, pid)
                owned[ev.team_id].add(aliased)
                if ev.role is not Role.GK:
                    orig_event_by_code[ev.team_id][aliased] = ev
            if ev.role is Role.GK:
                gk_block_event[ev.team_id] = (ev.event_id, ev.amount)
    remaining_g2 = {
        tid: t.budgets["G2"].remaining
        for tid, t in eff_state.teams.items()
        if "G2" in t.budgets
    }

    new_events: list = []
    errors: list[str] = []
    per_team_report: list[str] = []

    for team_name, slots in staged.teams.items():
        tid = _resolve_team_id(team_name, labels)
        if tid is None:
            errors.append(f"team not resolved: {team_name!r}")
            continue
        if args.only_team and tid != args.only_team:
            continue

        resolved: dict[str, str] = {}  # code -> role_letter (scoring role, listone)
        roster_gk: list[str] = []
        pos_roles = _positional_roles(len(slots), ruleset)
        slot_role_by_code: dict[str, Role] = {}  # code -> positional roster slot
        for i, slot in enumerate(slots):
            r = _resolve_player(slot.display_name, idx)
            if r is None:
                errors.append(f"[{team_name}] unresolved: {slot.display_name!r} ({slot.cost}cr)")
                continue
            code, role = r
            resolved[code] = role
            slot_role_by_code[code] = pos_roles[i]
            if role == "P":
                roster_gk.append(code)

        team_owned = owned.get(tid, set())
        team_new: list = []
        g3_spend = 0

        # 1. GK block: void + re-add if the roster's keepers differ from the G1 block
        ledger_gk = {p for p in team_owned if resolved.get(p) == "P" or p in roster_gk}
        # accurate ledger GK set: the codes in the G1 GK block event
        g1_gk_ids = set()
        for ev in existing:
            if isinstance(ev, AssignmentEvent) and ev.team_id == tid and ev.role is Role.GK:
                g1_gk_ids = {NEGATIVE_ALIAS.get(p, p) for p in ev.item.player_ids}
        gk_changed = set(roster_gk) != g1_gk_ids and tid in gk_block_event
        if gk_changed:
            old_id, old_amount = gk_block_event[tid]
            team_new.append(VoidEvent(
                event_id=uuid.uuid4().hex, ts=_TS, voids=old_id, author=AUTHOR,
                reason="lega_roster_reconciliation: blocco portieri cambiato in G3/G4",
            ))
            team_new.append(AssignmentEvent(
                event_id=uuid.uuid4().hex, ts=_TS, round_id=ROUND_ID, team_id=tid,
                pool_id=POOL_ID, role=Role.GK,
                item=AssignmentItem(player_ids=tuple(sorted(roster_gk))),
                amount=int(old_amount), source=SOURCE, author=AUTHOR, corrects=None,
            ))
            g3_spend += int(old_amount)
            # a re-added GK block does NOT consume fresh G3 budget beyond what G1
            # already paid; net it back out with a negative adjustment below.
            team_new.append(BudgetAdjustmentEvent(
                event_id=uuid.uuid4().hex, ts=_TS, round_id=ROUND_ID, team_id=tid,
                amount=int(old_amount),
                reason="lega_roster_reconciliation: rimborso costo blocco portieri G1 riassegnato in G3",
                author=AUTHOR,
            ))

        # 1b. owned outfield players sitting in a slot of a different role than
        # their scoring role (Isaksen / Rodriguez Je. type, ADR-2026-068/082):
        # supersede the original G1/G2 event so replay counts the positional
        # slot, not the scoring role, against the per-role cap. The correction is
        # recorded in G3 (that is when the roster actually settled, and `replay`
        # forbids a G1/G2 event appended after G3 ones already exist); `corrects`
        # drops the original from the current view, so its cost moves from its
        # round into G3. We bump this team's `remaining_G2` by the freed amount
        # so `g3_avail` rises exactly as `g3_spend` does -> gap and total budget
        # unchanged.
        g2_refund_from_corrections = 0
        for code, orig in orig_event_by_code.get(tid, {}).items():
            slot_role = slot_role_by_code.get(code)
            if slot_role is None or slot_role is Role.GK or slot_role is orig.role:
                continue
            g2_refund_from_corrections += int(orig.amount)
            g3_spend += int(orig.amount)
            team_new.append(AssignmentEvent(
                event_id=uuid.uuid4().hex, ts=_TS, round_id=ROUND_ID, team_id=tid,
                pool_id=POOL_ID, role=orig.role,
                item=AssignmentItem(player_ids=tuple(orig.item.player_ids)),
                amount=int(orig.amount), source=SOURCE, author=AUTHOR,
                corrects=orig.event_id, slot_role=slot_role,
            ))

        # 1c. released players: owned outfield (non-GK) codes no longer on the
        # team's roster in the export -> the team svincolo'd them before
        # riparazione. A VoidEvent drops the original assignment from the current
        # view: the player leaves the roster and the credits paid come back
        # (admin rule 2026-09-02: "recuperate la spesa fatta"). Keeper-block
        # swaps are handled in step 1, not here. `*` players are still on the
        # roster (kept, flagged left_serie_a) so they never reach this branch.
        released = [
            (code, orig)
            for code, orig in orig_event_by_code.get(tid, {}).items()
            if code not in resolved
        ]
        if len(released) > ruleset.max_releases_per_team:
            errors.append(
                f"[{team_name}] {len(released)} svincoli > max "
                f"{ruleset.max_releases_per_team}: {[c for c, _ in released]}"
            )
        for code, orig in released:
            g2_refund_from_corrections += int(orig.amount)  # frees budget into remaining_G2
            team_new.append(VoidEvent(
                event_id=uuid.uuid4().hex, ts=_TS, voids=orig.event_id, author=AUTHOR,
                reason="lega_roster_reconciliation: giocatore svincolato prima della riparazione "
                       f"(rimborso {orig.amount}cr, spesa pagata)",
            ))

        # 2. new outfield players
        cost_by_code = {}
        for slot in slots:
            r = _resolve_player(slot.display_name, idx)
            if r:
                cost_by_code[r[0]] = int(slot.cost)
        for code, role in resolved.items():
            if role == "P":
                continue
            if code in team_owned:
                continue
            cost = cost_by_code.get(code, 1)
            g3_spend += cost
            scoring_role = _ROLE_MAP[role]
            slot_role = slot_role_by_code.get(code)
            team_new.append(AssignmentEvent(
                event_id=uuid.uuid4().hex, ts=_TS, round_id=ROUND_ID, team_id=tid,
                pool_id=POOL_ID, role=scoring_role,
                item=AssignmentItem(player_ids=(code,)),
                amount=cost, source=SOURCE, author=AUTHOR, corrects=None,
                slot_role=slot_role if slot_role is not None and slot_role is not scoring_role else None,
            ))

        # 3. budget-gap adjustment (the honest total spend exceeds remaining_G2 + 40)
        g3_avail = evaluate_budget_expr(
            next(r for r in ruleset.rounds if r.id == "G3").budget_available_expr,
            {"G2": remaining_g2.get(tid, 0) + g2_refund_from_corrections},
        )
        gap = g3_spend - g3_avail
        if gap > 0:
            team_new.insert(0, BudgetAdjustmentEvent(
                event_id=uuid.uuid4().hex, ts=_TS, round_id=ROUND_ID, team_id=tid,
                amount=int(gap),
                reason=f"lega_roster_reconciliation: spesa reale rosa supera remaining_G2+40 di {gap} "
                       "(discrepanza budget lega, questione admin aperta)",
                author=AUTHOR,
            ))

        new_events.extend(team_new)
        n_ass = sum(1 for e in team_new if isinstance(e, AssignmentEvent))
        per_team_report.append(
            f"  {tid:8} {team_name:24} nuovi={n_ass:2} spesa_G3={g3_spend:3} "
            f"avail_G3={g3_avail:3} gap={max(gap,0):+d} gk_cambiato={gk_changed}"
        )

    print(f"Squadre: {len(staged.teams)}. Eventi nuovi: {len(new_events)} "
          f"({sum(1 for e in new_events if isinstance(e, AssignmentEvent))} assegnazioni, "
          f"{sum(1 for e in new_events if isinstance(e, VoidEvent))} void, "
          f"{sum(1 for e in new_events if isinstance(e, BudgetAdjustmentEvent))} agg. budget). "
          f"Errori: {len(errors)}.")
    for line in per_team_report:
        print(line)
    if errors:
        print("\nERRORI (nessuna scrittura):")
        for e in errors:
            print("  -", e)
        return 1

    # order: VoidEvents first (no round), then G3 events; keeps round-ordering happy
    new_events.sort(key=lambda e: 0 if isinstance(e, VoidEvent) else 1)
    try:
        state = replay(ruleset, effective_events(existing + new_events))
    except (DomainError, ConfigError) as exc:
        print(f"\nREPLAY (effective) FALLITO, nessuna scrittura: {exc}")
        return 1
    print("replay(effective_events(existing + new)) OK: invarianti rispettati.")
    for tid in sorted(state.teams):
        t = state.team(tid)
        n = sum(len(t.roster[r]) for r in Role)
        print(f"  {tid}: rosa {n} giocatori "
              f"(P{len(t.roster[Role.GK])} D{len(t.roster[Role.DEF])} "
              f"C{len(t.roster[Role.MID])} A{len(t.roster[Role.FWD])})")

    if not args.yes:
        print("\nDRY-RUN: nessuna scrittura. Rilancia con --yes per scrivere sul ledger.")
        return 0

    if args.relabel_team05:
        for tid, lab in labels.items():
            if _squash(lab) == _squash(TEAM05_OLD_LABEL):
                team_labels_store.set_label(labels_conn, tid, TEAM05_NEW_LABEL)
                print(f"relabel {tid}: {lab!r} -> {TEAM05_NEW_LABEL!r}")
    for ev in new_events:
        ledger_store.append_event(ledger_conn, ev)
    print(f"Scritti {len(new_events)} eventi. Ledger ora: {len(ledger_store.load_events(ledger_conn))} eventi.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
