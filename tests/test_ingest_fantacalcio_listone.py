import openpyxl
import pytest

from fantacalcio.ingest.fantacalcio_listone import (
    ListoneParseError,
    parse_quotazioni_file,
    parse_statistiche_file,
)


def _write_quotazioni_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutti"
    ws.append(["Quotazioni Fantacalcio Stagione 2025 26"])
    ws.append(["Id", "R", "RM", "Nome", "Squadra", "Qt.A", "Qt.I", "Diff.", "Qt.A M", "Qt.I M", "Diff.M", "FVM", "FVM M"])
    ws.append([4431, "P", "Por", "Carnesecchi", "Atalanta", 18, 14, 4, 18, 14, 4, 80, 80])
    wb.save(path)


def _write_statistiche_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutti"
    ws.append(["Statistiche Fantacalcio Stagione 2025 26"])
    ws.append(["Id", "R", "Rm", "Nome", "Squadra", "Pv", "Mv", "Fm", "Gf", "Gs", "Rp", "Rc", "R+", "R-", "Ass", "Amm", "Esp", "Au"])
    ws.append([4431, "P", "Por", "Carnesecchi", "Atalanta", 37, 6.36, 5.58, 0, 35, 2, 0, 0, 0, 0, 0, 0, 0])
    wb.save(path)


def test_parse_quotazioni_file(tmp_path):
    path = tmp_path / "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx"
    _write_quotazioni_xlsx(path)
    staged = parse_quotazioni_file(path)
    assert len(staged.frame) == 1
    assert staged.season_label == "2025_26"
    row = staged.frame.iloc[0]
    assert row["player_code"] == 4431
    assert row["team_name"] == "Atalanta"
    assert row["fvm_classic"] == 80


def test_parse_statistiche_file(tmp_path):
    path = tmp_path / "Statistiche_Fantacalcio_Stagione_2025_26.xlsx"
    _write_statistiche_xlsx(path)
    staged = parse_statistiche_file(path)
    assert staged.season_label == "2025_26"
    row = staged.frame.iloc[0]
    assert row["player_code"] == 4431
    assert row["matches_with_vote"] == 37
    assert row["fantamedia"] == 5.58


def test_explicit_season_label_overrides_filename(tmp_path):
    path = tmp_path / "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx"
    _write_quotazioni_xlsx(path)
    staged = parse_quotazioni_file(path, season_label="custom")
    assert staged.season_label == "custom"


def test_unparseable_filename_raises(tmp_path):
    path = tmp_path / "quotazioni.xlsx"
    _write_quotazioni_xlsx(path)
    with pytest.raises(ListoneParseError, match="does not match"):
        parse_quotazioni_file(path)


def test_missing_file_raises(tmp_path):
    with pytest.raises(ListoneParseError, match="not found"):
        parse_quotazioni_file(tmp_path / "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx")


def test_missing_column_raises(tmp_path):
    path = tmp_path / "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutti"
    ws.append(["Quotazioni Fantacalcio Stagione 2025 26"])
    ws.append(["Id", "R", "Nome"])  # missing most columns
    ws.append([4431, "P", "Carnesecchi"])
    wb.save(path)
    with pytest.raises(ListoneParseError, match="missing expected columns"):
        parse_quotazioni_file(path)


def test_same_player_code_across_quotazioni_and_statistiche(tmp_path):
    q_path = tmp_path / "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx"
    s_path = tmp_path / "Statistiche_Fantacalcio_Stagione_2025_26.xlsx"
    _write_quotazioni_xlsx(q_path)
    _write_statistiche_xlsx(s_path)
    q = parse_quotazioni_file(q_path)
    s = parse_statistiche_file(s_path)
    assert q.frame.iloc[0]["player_code"] == s.frame.iloc[0]["player_code"]
