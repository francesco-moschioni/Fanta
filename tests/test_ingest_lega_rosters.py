"""Parser-only tests for the league roster export grid (no identity resolution)."""

import openpyxl
import pytest

from fantacalcio.ingest.lega_rosters import LegaRosterParseError, parse_roster_file


def _write_grid(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROSE"
    # band 1: two teams side by side (name, costo, spacer) x2
    ws.append(["Team Alpha", "costo", None, "Team Beta", "costo", None])
    ws.append(["Carnesecchi", 40, None, "Svilar", 35, None])
    ws.append(["Bastoni", 65, None, "Dimarco *", 1, None])
    ws.append(["totale", 105, None, "totale", 36, None])
    wb.save(path)


def test_parses_two_teams_and_star_marker(tmp_path):
    p = tmp_path / "rosters.xlsx"
    _write_grid(p)
    staged = parse_roster_file(p)

    assert set(staged.teams) == {"Team Alpha", "Team Beta"}
    alpha = staged.teams["Team Alpha"]
    assert [s.clean_name for s in alpha] == ["Carnesecchi", "Bastoni"]
    assert [s.cost for s in alpha] == [40, 65]
    assert all(not s.left_serie_a for s in alpha)

    beta = staged.teams["Team Beta"]
    assert beta[1].clean_name == "Dimarco"
    assert beta[1].left_serie_a is True
    assert beta[1].display_name == "Dimarco *"
    assert staged.file_sha256  # populated


def test_missing_sheet_raises(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "NotRose"
    p = tmp_path / "x.xlsx"
    wb.save(p)
    with pytest.raises(LegaRosterParseError, match="Sheet 'ROSE'"):
        parse_roster_file(p)


def test_non_integer_cost_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROSE"
    ws.append(["Team Alpha", "costo", None])
    ws.append(["Carnesecchi", "quaranta", None])
    p = tmp_path / "bad.xlsx"
    wb.save(p)
    with pytest.raises(LegaRosterParseError, match="non-integer cost"):
        parse_roster_file(p)


def test_missing_file_raises(tmp_path):
    with pytest.raises(LegaRosterParseError, match="File not found"):
        parse_roster_file(tmp_path / "nope.xlsx")
