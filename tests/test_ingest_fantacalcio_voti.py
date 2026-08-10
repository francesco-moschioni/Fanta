import openpyxl
import pytest

from fantacalcio.ingest.fantacalcio_voti import (
    PANELS,
    VotiParseError,
    parse_filename,
    parse_voti_file,
)

_HEADER = ["Cod.", "Ruolo", "Nome", "Voto", "Gf", "Gs", "Rp", "Rs", "Rf", "Au", "Amm", "Esp", "Ass"]


def _write_synthetic_voti_xlsx(path, voto_giornata_1="6", voto_giornata_2="6.5*"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for panel in PANELS:
        ws = wb.create_sheet(panel)
        ws.append([f"Voti {panel} 1a giornata di campionato"])
        ws.append(["Solo su www.fantacalcio.it i voti ufficiali..."])
        ws.append(["QUESTO FILE NON PUO' ESSERE RIPRODOTTO NE' PUBBLICATO"])
        ws.append(["E' DA CONSIDERARSI AD USO PERSONALE ESCLUSIVO"])
        ws.append(["Atalanta"])  # team banner row
        ws.append(_HEADER)
        ws.append([4, "P", "Sportiello", voto_giornata_1, 0, 1, 0, 0, 0, 0, 0, 0, 0])
        ws.append([554, "D", "Zappacosta", voto_giornata_2, 0, 0, 0, 0, 0, 0, 0, 0, 1])
        ws.append(["Bologna"])  # second team banner row, mid-sheet
        ws.append([22, "C", "De Roon", "6", 1, 0, 0, 0, 0, 0, 0, 0, 0])
    wb.save(path)


class TestParseFilename:
    def test_valid_filename(self):
        info = parse_filename("Voti_Fantacalcio_Stagione_2025_26_Giornata_38.xlsx")
        assert info.season_start_year == 2025
        assert info.season_end_year_suffix == 26
        assert info.matchday == 38
        assert info.season_label == "2025_26"

    def test_invalid_filename_raises(self):
        with pytest.raises(VotiParseError, match="does not match"):
            parse_filename("random_file.xlsx")


class TestParseVotiFile:
    def test_parses_all_panels_and_drops_banner_rows(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        _write_synthetic_voti_xlsx(path)
        staged = parse_voti_file(path)

        assert staged.season_label == "2025_26"
        assert staged.matchday == 1
        # 3 real players * 3 panels = 9 rows; the 2 team-banner rows per panel are dropped.
        assert len(staged.frame) == 9
        assert set(staged.frame["panel"]) == set(PANELS)
        assert set(staged.frame["player_code"]) == {4, 554, 22}

    def test_provisional_voto_flagged_and_stripped(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        _write_synthetic_voti_xlsx(path)
        staged = parse_voti_file(path)

        zappacosta = staged.frame[staged.frame["player_code"] == 554]
        assert (zappacosta["voto"] == 6.5).all()
        assert zappacosta["voto_provisional"].all()

        sportiello = staged.frame[staged.frame["player_code"] == 4]
        assert (sportiello["voto"] == 6.0).all()
        assert not sportiello["voto_provisional"].any()

    def test_explicit_season_matchday_override_filename(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        _write_synthetic_voti_xlsx(path)
        staged = parse_voti_file(path, season_label="custom_label", matchday=99)
        assert staged.season_label == "custom_label"
        assert staged.matchday == 99

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(VotiParseError, match="not found"):
            parse_voti_file(tmp_path / "missing.xlsx")

    def test_unparseable_voto_raises(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        _write_synthetic_voti_xlsx(path, voto_giornata_1="SV")  # senza voto, not numeric
        with pytest.raises(VotiParseError, match="not numeric"):
            parse_voti_file(path)

    def test_missing_sheet_raises(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Fantacalcio"
        for row in [["x"], ["x"], ["x"], ["x"], ["Atalanta"], _HEADER]:
            wb.active.append(row)
        wb.save(path)
        with pytest.raises(VotiParseError, match="not found"):
            parse_voti_file(path)

    def test_missing_column_raises(self, tmp_path):
        path = tmp_path / "Voti_Fantacalcio_Stagione_2025_26_Giornata_1.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        bad_header = [c for c in _HEADER if c != "Voto"]
        for panel in PANELS:
            ws = wb.create_sheet(panel)
            for row in [["x"], ["x"], ["x"], ["x"], ["Atalanta"], bad_header]:
                ws.append(row)
            ws.append([4, "P", "Sportiello", 0, 1, 0, 0, 0, 0, 0, 0])
        wb.save(path)
        with pytest.raises(VotiParseError, match="missing expected columns"):
            parse_voti_file(path)
