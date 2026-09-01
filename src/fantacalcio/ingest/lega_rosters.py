"""Parser for the league-platform roster export ("20lega-rosters-*.xlsx").

Manual-import only, no HTTP -- same policy as `fantacalcio_voti.py`. The export
is a single sheet "ROSE" laid out as a wide grid: repeated 3-column blocks
(name, "costo", spacer), one block per fantasy team, with team blocks stacked in
horizontal bands separated by header rows that contain the literal "costo".

Rows under each team header are ``(player_display_name, cost)`` until a blank
cell or a "totale" row. A trailing ``*`` on a name is the platform's marker for
a player who has left Serie A (sold abroad / out of the league) -- kept verbatim
so the caller can decide what to do with it.

This module only *reads structure*; it does not resolve player identity or team
identity (that is the importer's job, against the player table and
`team_labels`).
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

import openpyxl


class LegaRosterParseError(ValueError):
    pass


@dataclass(frozen=True)
class RosterSlot:
    display_name: str  # verbatim, including a trailing " *" if present
    cost: int
    left_serie_a: bool  # True when the source name carried the "*" marker

    @property
    def clean_name(self) -> str:
        return self.display_name.rstrip(" *").strip()


@dataclass(frozen=True)
class StagedLegaRosters:
    file_path: str
    file_sha256: str
    sheet: str
    teams: dict[str, tuple[RosterSlot, ...]]  # platform team name -> roster


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def parse_roster_file(path: str | Path, sheet: str = "ROSE") -> StagedLegaRosters:
    path = Path(path)
    if not path.is_file():
        raise LegaRosterParseError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise LegaRosterParseError(f"Sheet {sheet!r} not in {path.name} (has {wb.sheetnames})")
    ws = wb[sheet]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    ncol = ws.max_column
    wb.close()

    header_rows = [
        i
        for i, row in enumerate(grid)
        if any(str(c).strip().lower() == "costo" for c in row if c is not None)
    ]
    if not header_rows:
        raise LegaRosterParseError("No 'costo' header row found; layout not recognised")

    teams: dict[str, list[RosterSlot]] = {}
    for hr in header_rows:
        for j in range(0, ncol, 3):
            name_cell = grid[hr][j] if j < len(grid[hr]) else None
            if name_cell is None:
                continue
            team_name = str(name_cell).strip()
            if team_name == "" or team_name.lower() == "costo":
                continue

            slots: list[RosterSlot] = []
            for k in range(hr + 1, len(grid)):
                pn = grid[k][j] if j < len(grid[k]) else None
                if pn is None or str(pn).strip() == "":
                    break
                sval = str(pn).strip()
                if sval.lower() == "totale":
                    break
                raw_cost = grid[k][j + 1] if j + 1 < len(grid[k]) else None
                try:
                    cost = int(raw_cost)
                except (TypeError, ValueError):
                    raise LegaRosterParseError(
                        f"{team_name!r} row {k + 1}: non-integer cost {raw_cost!r} for {sval!r}"
                    )
                slots.append(
                    RosterSlot(
                        display_name=sval,
                        cost=cost,
                        left_serie_a=sval.endswith("*"),
                    )
                )
            if slots:
                if team_name in teams and len(teams[team_name]) >= len(slots):
                    # a duplicate/echo header column with no more data -- ignore
                    continue
                teams[team_name] = slots

    if not teams:
        raise LegaRosterParseError("Parsed zero teams")

    return StagedLegaRosters(
        file_path=str(path),
        file_sha256=_sha256(path),
        sheet=sheet,
        teams={name: tuple(slots) for name, slots in teams.items()},
    )
